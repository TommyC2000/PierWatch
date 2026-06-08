import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from src.excel_inspector import inspect_workbook


class InspectWorkbookTests(unittest.TestCase):
    def test_returns_sheet_names_dimensions_first_rows_and_header_guess(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "sample.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "River Stage"
            ws.append([None, None, None])
            ws.merge_cells("A2:C2")
            ws["A2"] = "Merged title"
            ws.append([None, None, None])
            ws.append(["Date / Time", "Stage (Ft)", "Notes"])
            ws.append(["2026-01-01", 8.2, "low"])
            wb.save(path)

            info = inspect_workbook(path)

        self.assertEqual(info["sheet_names"], ["River Stage"])
        sheet = info["sheets"]["River Stage"]
        self.assertEqual(sheet["dimensions"], {"max_row": 5, "max_column": 3})
        self.assertEqual(sheet["first_rows"][1]["values"], ["Merged title", "Merged title", "Merged title"])
        self.assertEqual(sheet["guessed_header_row"], 4)

    def test_empty_sheet_does_not_guess_header(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "empty.xlsx"
            wb = Workbook()
            wb.active.title = "Empty"
            wb.save(path)

            info = inspect_workbook(path)

        sheet = info["sheets"]["Empty"]
        self.assertIsNone(sheet["guessed_header_row"])
        self.assertEqual(sheet["first_rows"][0]["values"], [None])


if __name__ == "__main__":
    unittest.main()
