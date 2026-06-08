from __future__ import annotations

from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


HEADER_KEYWORDS = (
    "date",
    "time",
    "stage",
    "gage",
    "gps",
    "long",
    "trans",
    "temperature",
    "expansion",
    "corrected",
    "pier",
    "joint",
)


def _serialize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _build_merged_lookup(ws, max_row: int, max_column: int) -> dict[tuple[int, int], Any]:
    lookup: dict[tuple[int, int], Any] = {}
    merged_cells = getattr(ws, "merged_cells", None)
    if not merged_cells:
        return lookup

    for cell_range in merged_cells.ranges:
        if cell_range.min_row > max_row or cell_range.min_col > max_column:
            continue
        value = ws.cell(cell_range.min_row, cell_range.min_col).value
        for row_idx in range(cell_range.min_row, min(cell_range.max_row, max_row) + 1):
            for col_idx in range(cell_range.min_col, min(cell_range.max_col, max_column) + 1):
                lookup[(row_idx, col_idx)] = value
    return lookup


def _read_row(ws, row_idx: int, max_column: int, merged_lookup: dict[tuple[int, int], Any]) -> list[Any]:
    values = []
    for col_idx in range(1, max_column + 1):
        value = ws.cell(row=row_idx, column=col_idx).value
        if value is None and (row_idx, col_idx) in merged_lookup:
            value = merged_lookup[(row_idx, col_idx)]
        values.append(_serialize_cell(value))
    return values


def _guess_header_row(rows: list[dict[str, Any]]) -> int | None:
    best_row = None
    best_score = 0.0

    for row in rows:
        values = row["values"]
        non_blank_values = [value for value in values if not _is_blank(value)]
        if not non_blank_values:
            continue

        row_text = " ".join(str(value).lower() for value in non_blank_values)
        keyword_hits = sum(1 for keyword in HEADER_KEYWORDS if keyword in row_text)
        text_cells = sum(isinstance(value, str) for value in non_blank_values)
        numeric_cells = sum(isinstance(value, (int, float)) and not isinstance(value, bool) for value in non_blank_values)

        score = len(non_blank_values) + (3 * keyword_hits) + min(text_cells, 4) - (0.25 * numeric_cells)
        if keyword_hits == 0 and len(non_blank_values) < 2:
            score -= 2

        if score > best_score:
            best_score = score
            best_row = row["row_number"]

    return best_row if best_score >= 2 else None


def inspect_workbook(excel_path: str | Path, sample_rows: int = 10, header_scan_rows: int = 30, max_sample_columns: int = 60) -> dict:
    """Inspect workbook shape and early rows without parsing engineering data."""
    path = Path(excel_path)
    result = {
        "workbook_path": str(path),
        "sheet_names": [],
        "sheets": {},
    }

    try:
        wb = load_workbook(path, read_only=False, data_only=True)
    except Exception as exc:
        result["workbook_error"] = str(exc)
        return result

    result["sheet_names"] = list(wb.sheetnames)

    for ws in wb.worksheets:
        try:
            max_row = ws.max_row or 0
            max_column = ws.max_column or 0
            sample_column_count = min(max(max_column, 1), max_sample_columns)
            scan_row_count = min(max(max_row, 1), max(sample_rows, header_scan_rows))
            merged_lookup = _build_merged_lookup(ws, scan_row_count, sample_column_count)

            scanned_rows = [
                {
                    "row_number": row_idx,
                    "values": _read_row(ws, row_idx, sample_column_count, merged_lookup),
                    "non_empty_count": None,
                }
                for row_idx in range(1, scan_row_count + 1)
            ]
            for row in scanned_rows:
                row["non_empty_count"] = sum(not _is_blank(value) for value in row["values"])

            result["sheets"][ws.title] = {
                "sheet_name": ws.title,
                "dimensions": {"max_row": max_row, "max_column": max_column},
                "max_row": max_row,
                "max_column": max_column,
                "first_rows": scanned_rows[:sample_rows],
                "guessed_header_row": _guess_header_row(scanned_rows),
                "merged_range_count": len(getattr(ws.merged_cells, "ranges", [])),
                "sampled_column_count": sample_column_count,
                "truncated_columns": max_column > sample_column_count,
            }
        except Exception as exc:
            result["sheets"][ws.title] = {
                "sheet_name": ws.title,
                "dimensions": {"max_row": ws.max_row or 0, "max_column": ws.max_column or 0},
                "max_row": ws.max_row or 0,
                "max_column": ws.max_column or 0,
                "first_rows": [],
                "guessed_header_row": None,
                "inspection_error": str(exc),
            }

    return result


def list_sheets(excel_path: str) -> list[str]:
    return pd.ExcelFile(excel_path).sheet_names
